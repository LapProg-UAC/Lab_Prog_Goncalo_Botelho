import random
import openpyxl
import json
import encriptacao

CHAVE = [3, 5, 2, 9]

def abrirficheiromedicamentos(caminho: str):
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            conteudo = f.read().replace("", "")
            return [l.strip() for l in conteudo.splitlines() if l.strip()]
    except FileNotFoundError:
        print("Erro: O ficheiro", caminho, "não foi encontrado.")
        return []

def criartabelasinoptica(medicamentos: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabela Sinóptica"

    RegrasInteracao = {}

    for i, Nome_i in enumerate(medicamentos, start=2):
        ws.cell(row=i, column=1, value=Nome_i)
        ws.cell(row=1, column=i, value=Nome_i)

        for j, Nome_j in enumerate(medicamentos, start=2):
            if i == j:
                Valor = 0
            else:
                Valor = random.randint(0, 6)

            ws.cell(row=i, column=j, value=Valor)
            RegrasInteracao[(Nome_i, Nome_j)] = Valor

    return wb, RegrasInteracao

def salvartabelasinoptica(nomeficheiro: str, wb: openpyxl.Workbook):
    try:
        wb.save(nomeficheiro + ".xlsx")

        with open(nomeficheiro + ".txt", "w", encoding="utf-8") as f:
            for row in wb.active.iter_rows(values_only=True):
                f.write("\t".join(str(cell) if cell is not None else "" for cell in row) + "\n")

        print(f"Tabela salva como {nomeficheiro}.xlsx e .txt")

    except Exception as e:
        print("Erro ao salvar:", e)

def gerardadosfinais(caminhonomes: str, caminhosobrenomes: str, numutentes: int, listamedicamentos: list, Regras: dict, CHAVE):
    NomesBase = abrirficheiromedicamentos(caminhonomes)
    SobrenomesBase = abrirficheiromedicamentos(caminhosobrenomes)

    ListaIDs = []
    ListaJson = []

    # Gerar IDs únicos
    i = 0
    while i < numutentes:
        NumeroID = random.randint(1000, 9999)
        if NumeroID not in ListaIDs:
            ListaIDs.append(NumeroID)
            i += 1

    for uid in ListaIDs:
        NomeCompleto = f"{random.choice(NomesBase)} {random.choice(SobrenomesBase)}"
        Receita = random.sample(listamedicamentos, k=random.randint(3, 5))

        SomaInteracoes = 0
        Interage = False

        for idx in range(len(Receita)):
            for j in range(idx + 1, len(Receita)):
                Grau = Regras.get((Receita[idx], Receita[j]), 0)
                SomaInteracoes += Grau
                if Grau > 0:
                    Interage = True

        # Aqui o número de utente (uid) é convertido para string e encriptado usando a chave
        id_encriptado = encriptacao.encriptar(str(uid), CHAVE)

        ListaJson.append({
            "id_utente": id_encriptado,
            "nome_utente": NomeCompleto,
            "receita": Receita,
            "balanco_total": SomaInteracoes,
            "resultado": "Com Interações" if Interage else "Sem Interações"
        })

    return ListaJson

def salvarjson(dados: list, nome: str):
    with open(nome, "w", encoding="utf-8") as f:
        json.dump(dados, f, indent=4, ensure_ascii=False)

    print(f"Ficheiro JSON '{nome}' criado.")

def main():
    listamedicamentos = abrirficheiromedicamentos("medicamentos.txt")
    wb, Regras = criartabelasinoptica(listamedicamentos)
    salvartabelasinoptica("tabela_sinoptica", wb)

    DadosFinais = gerardadosfinais("nomes.txt", "sobrenomes.txt", 100, listamedicamentos, Regras, CHAVE)
    salvarjson(DadosFinais, "dados_final.json")

if __name__ == "__main__":  main()
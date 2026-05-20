# --- Imports ---
import os
import openpyxl
import random
import sys
from typing import List


# --- Constants ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

INPUT_DIR = os.path.join(BASE_DIR, "InputFiles")
OUTPUT_DIR = os.path.join(BASE_DIR, "OutputFiles")

os.makedirs(OUTPUT_DIR, exist_ok=True)

MEDS_FILE = os.path.join(INPUT_DIR, "medications.txt") 
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "sinoptic_table.xlsx")

MIN_VALUE: int = 0
MAX_VALUE: int = 5

# --- Functions ---
def ReadMeds(FileDirectory: str, Type: str = "utf-8") -> List[str]:
    """
    Reads the medication file and returns a list of medications ignoring empty lines.

    :param FileDirectory(str): Path to the medication file.
    :param Type(str): Type of the file.
    :return(List[str]): List of medications.
    """
    try:
        with open(FileDirectory, "r", encoding=Type) as File:
            return [Line.strip() for Line in File if Line.strip()]
        
    except FileNotFoundError:
        print(f"Error: The file '{FileDirectory}' was not found.")
        sys.exit(1)


def GenerateSinopticTable(Meds: List[str]) -> List[List[int]]:
    """
    Generates the square matrix with 0 on the diagonal.

    :param Meds(List[str]): List of medications.
    :return(List[List[int]]): Interaction matrix.
    """
    n = len(Meds)
    
    return [[0 if i == j else random.randint(MIN_VALUE, MAX_VALUE) for j in range(n)] for i in range(n)]


def SaveToExcel(Meds: List[str], Table: List[List[int]], OutputFile: str) -> None:
    """
    Saves the table to a native Excel file (.xlsx).

    :param Meds(List[str]): List of medications.
    :param Table(List[List[int]]): Interaction matrix.
    :param OutputFile(str): Path to the output file.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Interactions"

        for col_idx, nome in enumerate(Meds, start=2):
            ws.cell(row=1, column=col_idx, value=nome)

        for RowIndex, MedName in enumerate(Meds, start=2):
            ws.cell(row=RowIndex, column=1, value=MedName)
            for ColIndex, Value in enumerate(Table[RowIndex-2], start=2):
                ws.cell(row=RowIndex, column=ColIndex, value=Value)

        wb.save(OutputFile)
        print(f"[OK] File Excel '{OutputFile}' generated successfully.")
        
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        sys.exit(1)

# --- Main ---
def Main():
    """
    Main function that orchestrates the reading of medications, generation of the sinoptic table and saving of the Excel file.

    - Reads the medications from the text file.
    - Generates the sinoptic table with random interactions.
    - Saves the table to a native Excel file (.xlsx).
    - Includes error handling for files not found and issues with Excel writing.
    - Maintains a modular and clear structure for easy maintenance and future extension.
    """
    Meds = ReadMeds(MEDS_FILE)
    if not Meds: return
    
    SinopticTable = GenerateSinopticTable(Meds)
    SaveToExcel(Meds, SinopticTable, OUTPUT_FILE)


if __name__ == "__main__":
    Main()
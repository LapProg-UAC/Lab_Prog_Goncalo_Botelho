# --- Imports ---
import os
import openpyxl
import json
import random
import sys
from typing import Dict, List, Tuple


# --- Constants ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

INPUT_DIR = os.path.join(BASE_DIR, "InputFiles")
OUTPUT_DIR = os.path.join(BASE_DIR, "OutputFiles")

EXCEL_FILE = os.path.join(OUTPUT_DIR, "sinoptic_table.xlsx")
NAMES_FILE = os.path.join(INPUT_DIR, "names.txt")
SURNAMES_FILE = os.path.join(INPUT_DIR, "surnames.txt")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "prescriptions.json")

PATIENTS_NUM: int = 10
MIN_MEDS: int = 2
MAX_MEDS: int = 6
INTERACTION_LIMIT: int = 1

INTERACTIONS_MAPPING: Dict[int, str] = {
    0: "No clinical significance",
    1: "Potentially serious",
    2: "Enhancer of therapeutic/toxic effect (horizontal column)",
    3: "Enhancer of therapeutic/toxic effect (vertical column)",
    4: "Decreases therapeutic/toxic effect (horizontal column)",
    5: "Decreases therapeutic/toxic effect (vertical column)",
}


# --- Aux ---
SinopticTable = Dict[str, Dict[str, int]]
Patient = Dict[str, str]
Prescription = Dict[str, List[str]]

def NormalizeName(Name: str) -> str:
    """
    Normalizes a name by stripping whitespace and converting to lowercase.

    :param Name: Name to normalize.
    :return: Normalized name.
    """
    if not Name: 
        return ""
    return str(Name).strip().lower()


# --- Funções de leitura ---
def ReadFile(Path: str, Type: str = "utf-8") -> List[str]:
    """
    Reads a list from a text file, ignoring empty lines.

    :param Path: Path to the file.
    :return: List of strings read from the file.
    """
    try:
        with open(Path, "r", encoding=Type) as f:
            return [Line.strip() for Line in f if Line.strip()]
        
    except FileNotFoundError:
        print(f"Error: File '{Path}' not found.")
        sys.exit(1)


def ReadExcel(Path: str) -> Tuple[List[str], SinopticTable]:
    """
    Reads the Excel table using openpyxl.

    :param Path: Path to the Excel file.
    :return: Tuple containing the list of names and the sinoptic table.
    """
    try:
        wb = openpyxl.load_workbook(Path, data_only=True)
        ws = wb.active

        OriginalNames = []
        for col in range(2, ws.max_column + 1):
            Value = ws.cell(row=1, column=col).value
            if Value:
                OriginalNames.append(str(Value))

        Table = {}
        for row in range(2, ws.max_row + 1):
            LineName = ws.cell(row=row, column=1).value
            if not LineName: continue
            
            KeyLine = NormalizeName(LineName)
            Table[KeyLine] = {}

            for col_idx, name_col in enumerate(OriginalNames, start=2):
                KeyColumn = NormalizeName(name_col)
                ValueCell = ws.cell(row=row, column=col_idx).value
                
                try:
                    Table[KeyLine][KeyColumn] = int(ValueCell) if ValueCell is not None else 0
                except (ValueError, TypeError):
                    Table[KeyLine][KeyColumn] = 0
        return OriginalNames, Table
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        sys.exit(1)


# --- Lógica de Geração e Cálculo ---
def CalculateBalance(Prescription: List[str], Table: SinopticTable) -> Dict:
    """
    Calculates the interaction balance for a given prescription based on the sinoptic table.

    :param Prescription: List of prescribed medications.
    :param Table: Sinoptic table of interactions.
    :return: Dictionary with the interaction balance.
    """
    Interactions = []
    Count = {v: 0 for v in range(6)}
    
    n = len(Prescription)
    for i in range(n):
        for j in range(i + 1, n):
            med_a = Prescription[i]
            med_b = Prescription[j]
            
            v = Table.get(NormalizeName(med_a), {}).get(NormalizeName(med_b), 0)
            Count[v] += 1
            Interactions.append({
                "Medicine_a": med_a,
                "Medicine_b": med_b,
                "Value": v,
                "Description": INTERACTIONS_MAPPING.get(v, "Unknown")
            })

    Have_Interactions = any(p["Value"] >= INTERACTION_LIMIT for p in Interactions)
    return {
        "Analyzed_Pairs": len(Interactions),
        "Interactions": sorted(Interactions, key=lambda x: x["Value"], reverse=True),
        "Count_by_Type": {str(k): v for k, v in Count.items()},
        "Have_Interactions": Have_Interactions,
        "Decision": "With Interactions" if Have_Interactions else "Without Interactions"
    }


def Main(Type: str = "utf-8") -> None:
    """
    Main function that reads the data, generates the prescriptions and calculation of the interaction balance.
    
    - Reads the synoptic table from the Excel file.
    - Reads user's first and last names.
    - Generates random recipes for a defined number of users.
    - It calculates the balance of interactions for each recipe.
    """
    Meds, Table = ReadExcel(EXCEL_FILE)
    Names = ReadFile(NAMES_FILE)
    Surnames = ReadFile(SURNAMES_FILE)

    Prescription = []
    for i in range(PATIENTS_NUM):
        PatientName = f"{random.choice(Names)} {random.choice(Surnames)}"
        MedsNum = random.randint(MIN_MEDS, min(MAX_MEDS, len(Meds)))
        MedsList = random.sample(Meds, MedsNum)
        
        Balance = CalculateBalance(MedsList, Table)
        
        Prescription.append({
            "Patient_ID": f"U-{i+1:05d}",
            "Patient_Name": PatientName,
            "Prescribed_Medications": MedsList,
            "Num_Medicines": len(MedsList),
            "Balance_Interactions": Balance
        })

    with open(OUTPUT_FILE, "w", encoding=Type) as f:
        json.dump({"Total_Patients": len(Prescription), "Prescription": Prescription}, f, ensure_ascii=False, indent=4)
    
    print(f"[OK] Generated {len(Prescription)} Prescription in '{OUTPUT_FILE}'.")


if __name__ == "__main__":
    Main()